import openpyxl

try:
    wb = openpyxl.load_workbook(r"d:\Information management system\BoK.xlsx", data_only=True)
    for name in wb.sheetnames:
        print(f"\n--- Sheet: {name} ---")
        sheet = wb[name]
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx <= 25:
                print(f"Row {row_idx}: {row}")
except Exception as e:
    print("Error:", e)
